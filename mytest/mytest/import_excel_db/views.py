from http.client import HTTPResponse

import openpyxl
from django.http import StreamingHttpResponse, HttpResponse, Http404
from django.shortcuts import render, redirect
import pandas as pd
import os
from django.core.files.storage import FileSystemStorage
from openpyxl.packaging.manifest import mimetypes
from openpyxl.styles import PatternFill

from mytest import settings
from .models import Employee
from tablib import Dataset
from .templates.import_excel_db.resources import EmployeeResource


def Import_Excel_pandas(request):

    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        fs = FileSystemStorage()
        filename = fs.save(myfile.name, myfile)
        empexceldata = pd.read_excel(filename)
        dbframe = empexceldata

        # data_to_be_deleted = Employee.objects.all()
        # data_to_be_deleted.delete()

        path = settings.BASE_DIR
        os.remove(os.path.join(path, filename))

        def returnNotNull(frame):
            if frame != frame:
                return 0
            else:
                return frame

        def CalculateE(param):
            if param < 5000000:
                return param*0.13
            else:
                return param*0.15

        def CheckColor(param):
            return param == 0

        filename = 'example.xlsx'
        path = settings.BASE_DIR
        filepath = os.path.join(path, filename)
        book = openpyxl.load_workbook(filename=filepath)
        sheet = book.active
        i_row = 3

        fillgreen = PatternFill('solid', fgColor='008000')
        fillred = PatternFill('solid', fgColor='ff0000')

        for dbframe in dbframe.itertuples(index=False):
            if isinstance(dbframe[0], str) and isinstance(dbframe[1], str):
                if dbframe[0] == 'Владивостокское представительство ТК РГ':
                    sheet.insert_rows(i_row)
                    sheet['A'+str(i_row)].value = dbframe[0]
                    sheet['B'+str(i_row)].value = dbframe[1]
                    sheet['C' + str(i_row)].value = returnNotNull(dbframe[4])
                    sheet['D' + str(i_row)].value = returnNotNull(dbframe[5])
                    sheet['E' + str(i_row)].value = CalculateE(returnNotNull(dbframe[4]))
                    newF = returnNotNull(dbframe[5])-CalculateE(returnNotNull(dbframe[4]))
                    sheet['F' + str(i_row)].value = newF
                    if CheckColor(newF):
                        sheet['F' + str(i_row)].fill = fillgreen
                    else:
                        sheet['F' + str(i_row)].fill = fillred
                    i_row += 1

                    # obj = Employee.objects.create(branch=dbframe[0],
                    #                             employee=dbframe[1],
                    #                             accrued=returnNotNull(dbframe[2]),
                    #                             deductions=returnNotNull(dbframe[3]),
                    #                             tax=returnNotNull(dbframe[4]),
                    #                             calculated=returnNotNull(dbframe[5]),
                    #                             withheld=returnNotNull(dbframe[6])
                    #                             )
                    # obj.save()
        sheet.auto_filter.ref = 'A1:F999'
        book.save('new.xlsx')
        return redirect('download')

    return render(request, 'import_excel_db/import_excel_db.html', {})

def download_file(request):
    file_path = os.path.join(settings.BASE_DIR, 'new.xlsx')
    if os.path.exists(file_path):
        with open(file_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
            return response
    raise Http404
